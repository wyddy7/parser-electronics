"""Запись результатов парсинга в Excel файл"""
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional, List
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import structlog


logger = structlog.get_logger()


class ExcelWriter:
    """Записывает результаты парсинга в новый Excel файл"""
    
    # Стили для форматирования
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    
    STATUS_FOUND_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Зеленый
    STATUS_NOT_FOUND_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Красный
    STATUS_ERROR_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Желтый
    
    def __init__(self, output_dir: str = "output"):
        """
        Инициализация writer.
        
        Args:
            output_dir: Директория для выходных файлов
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.log = logger.bind(module="ExcelWriter")
    
    def write_results(
        self,
        original_df: pd.DataFrame,
        results: Dict[str, Optional[Dict]],
        output_filename: Optional[str] = None,
        parser_name: Optional[str] = None
    ) -> str:
        """
        Записывает результаты парсинга в новый Excel файл.
        
        Args:
            original_df: Оригинальный DataFrame из входного файла
            results: Словарь с результатами парсинга {product_name: {price: ..., name: ..., url: ...}}
            output_filename: Имя выходного файла (если None - генерируется автоматически)
            
        Returns:
            Путь к созданному файлу
        """
        self.log.info("writing_results_started",
                     total_products=len(original_df),
                     results_count=len(results))
        
        # Создаем копию оригинального DataFrame
        result_df = original_df.copy()
        
        # Динамические названия колонок
        parser_display_name = parser_name.replace('_', ' ').title() if parser_name else "Parser"
        price_column = f'Цена {parser_display_name}'
        
        # Добавляем новые колонки
        result_df[price_column] = None
        result_df['Название на сайте'] = None
        result_df['Ссылка на товар'] = None
        result_df['Статус'] = 'Не обработан'
        
        # Заполняем результатами
        for idx, row in result_df.iterrows():
            product_name = row.get('product_name', '')
            
            if product_name in results:
                result = results[product_name]
                
                if result is not None:
                    price = result.get('price', 0)
                    result_df.at[idx, price_column] = price
                    
                    # Определяем статус по цене
                    if price is not None and price > 0:
                        # Товар найден с ценой - записываем название и ссылку
                        result_df.at[idx, 'Название на сайте'] = result.get('name')
                        result_df.at[idx, 'Ссылка на товар'] = result.get('url')
                        result_df.at[idx, 'Статус'] = 'Найдено'
                    elif price == -2:
                        # Цена по запросу
                        result_df.at[idx, 'Название на сайте'] = result.get('name')
                        result_df.at[idx, 'Ссылка на товар'] = result.get('url')
                        result_df.at[idx, 'Статус'] = 'Цена по запросу'
                    elif price == -1:
                        # Товар найден, но снят с производства
                        result_df.at[idx, 'Название на сайте'] = result.get('name')
                        result_df.at[idx, 'Ссылка на товар'] = result.get('url')
                        result_df.at[idx, 'Статус'] = 'Снят с производства'
                    elif price == 0:
                        # Товар НЕ найден - НЕ записываем название (оставляем пустым)
                        result_df.at[idx, 'Название на сайте'] = None
                        result_df.at[idx, 'Ссылка на товар'] = None
                        result_df.at[idx, 'Статус'] = 'Товар не найден на сайте'
                    else:
                        result_df.at[idx, 'Название на сайте'] = None
                        result_df.at[idx, 'Ссылка на товар'] = None
                        result_df.at[idx, 'Статус'] = 'Ошибка парсинга'
                else:
                    result_df.at[idx, 'Статус'] = 'Не найдено'
        
        # Генерируем имя файла если не указано
        if output_filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            parser_suffix = parser_name if parser_name else "parser"
            output_filename = f"prices_{parser_suffix}_{timestamp}.xlsx"
        
        output_path = self.output_dir / output_filename
        
        # Записываем в Excel
        result_df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Применяем форматирование
        self._apply_formatting(output_path, result_df)
        
        # Статистика
        found_count = (result_df['Статус'] == 'Найдено').sum()
        on_request_count = (result_df['Статус'] == 'Цена по запросу').sum()
        discontinued_count = (result_df['Статус'] == 'Снят с производства').sum()
        not_found_count = ((result_df['Статус'] == 'Товар не найден на сайте') | 
                          (result_df['Статус'] == 'Не найдено') |
                          (result_df['Статус'] == 'Не обработан')).sum()
        
        self.log.info("writing_results_completed",
                     output_file=str(output_path),
                     total=len(result_df),
                     found=found_count,
                     on_request=on_request_count,
                     discontinued=discontinued_count,
                     not_found=not_found_count)
        
        return str(output_path)
    
    def _apply_formatting(self, file_path: Path, df: pd.DataFrame):
        """
        Применяет форматирование к Excel файлу.
        
        Args:
            file_path: Путь к файлу
            df: DataFrame для определения размеров
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Форматируем заголовки
            for col_num in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Находим колонку со статусом
            status_col_idx = None
            for idx, col_name in enumerate(df.columns, 1):
                if col_name == 'Статус':
                    status_col_idx = idx
                    break
            
            # Применяем цветовую кодировку статусов
            if status_col_idx:
                for row_num in range(2, len(df) + 2):  # +2 потому что 1 строка заголовок, нумерация с 1
                    cell = ws.cell(row=row_num, column=status_col_idx)
                    status = cell.value
                    
                    if status == 'Найдено':
                        cell.fill = self.STATUS_FOUND_FILL
                    elif status == 'Не найдено':
                        cell.fill = self.STATUS_NOT_FOUND_FILL
                    elif status == 'Ошибка парсинга':
                        cell.fill = self.STATUS_ERROR_FILL
            
            # Автоподбор ширины колонок
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем изменения
            wb.save(file_path)
            wb.close()
            
            self.log.debug("formatting_applied", file=str(file_path))
            
        except Exception as e:
            self.log.warning("formatting_failed", error=str(e), error_type=type(e).__name__)
            # Не критическая ошибка - файл уже создан, просто без форматирования
    
    def create_summary(self, results: Dict[str, Optional[Dict]]) -> Dict[str, int]:
        """
        Создает сводку по результатам парсинга.
        
        Args:
            results: Словарь с результатами
            
        Returns:
            Словарь со статистикой
        """
        summary = {
            'total': len(results),
            'found': sum(1 for v in results.values() if v is not None),
            'not_found': sum(1 for v in results.values() if v is None),
        }
        
        self.log.info("summary_created", **summary)
        return summary
    
    def write_results_parallel(
        self,
        original_df: pd.DataFrame,
        results: Dict[str, Dict[str, Optional[Dict]]],
        output_filename: Optional[str] = None,
        parser_names: Optional[List[str]] = None
    ) -> str:
        """
        Записывает результаты парсинга от нескольких парсеров в параллельном режиме.
        
        Args:
            original_df: Оригинальный DataFrame из входного файла
            results: Словарь с результатами парсинга {product_name: {parser_name: {price: ..., name: ..., url: ...}}}
            output_filename: Имя выходного файла (если None - генерируется автоматически)
            parser_names: Список имен парсеров
            
        Returns:
            Путь к созданному файлу
        """
        if parser_names is None or len(parser_names) == 0:
            raise ValueError("parser_names должен содержать хотя бы один парсер")
        
        self.log.info("writing_results_parallel_started",
                     total_products=len(original_df),
                     results_count=len(results),
                     parsers=parser_names)
        
        # Создаем копию оригинального DataFrame
        result_df = original_df.copy()
        
        # Создаем колонки для каждого парсера
        for parser_name in parser_names:
            parser_display = parser_name.replace('_', ' ').title()
            
            result_df[f'Цена {parser_display}'] = None
            result_df[f'Название {parser_display}'] = None
            result_df[f'Ссылка {parser_display}'] = None
            result_df[f'Статус {parser_display}'] = 'Не обработан'
        
        # Общая колонка статуса (опционально, для удобства)
        result_df['Статус'] = 'Не обработан'
        
        # Опционально: колонка разницы цен (если 2 парсера)
        if len(parser_names) == 2:
            result_df['Разница цен'] = None
        
        # Заполняем данными
        for idx, row in result_df.iterrows():
            product_name = row.get('product_name', '')
            
            if product_name in results:
                product_results = results[product_name]
                
                prices = []
                for parser_name in parser_names:
                    parser_display = parser_name.replace('_', ' ').title()
                    result = product_results.get(parser_name) if product_results else None
                    
                    if result is not None:
                        price = result.get('price', 0)
                        result_df.at[idx, f'Цена {parser_display}'] = price
                        result_df.at[idx, f'Название {parser_display}'] = result.get('name')
                        result_df.at[idx, f'Ссылка {parser_display}'] = result.get('url')
                        
                        # Определяем статус для каждого парсера (как в одиночном режиме)
                        if price is not None and price > 0:
                            result_df.at[idx, f'Статус {parser_display}'] = 'Найдено'
                            prices.append(price)
                        elif price == -2:
                            result_df.at[idx, f'Статус {parser_display}'] = 'Цена по запросу'
                        elif price == -1:
                            result_df.at[idx, f'Статус {parser_display}'] = 'Снят с производства'
                        elif price == 0:
                            result_df.at[idx, f'Статус {parser_display}'] = 'Товар не найден на сайте'
                        else:
                            result_df.at[idx, f'Статус {parser_display}'] = 'Ошибка парсинга'
                    else:
                        # Парсер не нашел товар
                        result_df.at[idx, f'Цена {parser_display}'] = None
                        result_df.at[idx, f'Название {parser_display}'] = None
                        result_df.at[idx, f'Ссылка {parser_display}'] = None
                        result_df.at[idx, f'Статус {parser_display}'] = 'Не найдено'
                
                # Определяем общий статус (для удобства)
                if len(prices) == len(parser_names):
                    result_df.at[idx, 'Статус'] = 'Найдено во всех'
                elif len(prices) > 0:
                    result_df.at[idx, 'Статус'] = 'Найдено частично'
                else:
                    result_df.at[idx, 'Статус'] = 'Не найдено'
                
                # Разница цен (если 2 парсера и оба нашли цену)
                if len(parser_names) == 2 and len(prices) == 2:
                    result_df.at[idx, 'Разница цен'] = abs(prices[0] - prices[1])
        
        # Генерируем имя файла если не указано
        if output_filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            parser_suffix = '_'.join(parser_names)
            output_filename = f"prices_parallel_{parser_suffix}_{timestamp}.xlsx"
        
        output_path = self.output_dir / output_filename
        
        # Записываем в Excel
        result_df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Применяем форматирование
        self._apply_formatting_parallel(output_path, result_df, parser_names)
        
        # Статистика
        found_all_count = (result_df['Статус'] == 'Найдено во всех').sum()
        found_partial_count = (result_df['Статус'] == 'Найдено частично').sum()
        not_found_count = (result_df['Статус'] == 'Не найдено').sum()
        
        self.log.info("writing_results_parallel_completed",
                     output_file=str(output_path),
                     total=len(result_df),
                     found_all=found_all_count,
                     found_partial=found_partial_count,
                     not_found=not_found_count)
        
        return str(output_path)
    
    def _apply_formatting_parallel(self, file_path: Path, df: pd.DataFrame, parser_names: List[str]):
        """
        Применяет форматирование к Excel файлу для параллельного режима.
        
        Args:
            file_path: Путь к файлу
            df: DataFrame для определения размеров
            parser_names: Список имен парсеров
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Форматируем заголовки
            for col_num in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Находим колонки со статусами (общий и для каждого парсера)
            status_cols = {}
            for idx, col_name in enumerate(df.columns, 1):
                if col_name == 'Статус':
                    status_cols['общий'] = idx
                elif col_name.startswith('Статус '):
                    parser_name = col_name.replace('Статус ', '')
                    status_cols[parser_name] = idx
            
            # Применяем цветовую кодировку статусов для общего статуса
            if 'общий' in status_cols:
                status_col_idx = status_cols['общий']
                for row_num in range(2, len(df) + 2):
                    cell = ws.cell(row=row_num, column=status_col_idx)
                    status = cell.value
                    
                    if status == 'Найдено во всех':
                        cell.fill = self.STATUS_FOUND_FILL
                    elif status == 'Найдено частично':
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Светло-желтый
                    elif status == 'Не найдено':
                        cell.fill = self.STATUS_NOT_FOUND_FILL
            
            # Применяем цветовую кодировку статусов для каждого парсера
            for parser_name, status_col_idx in status_cols.items():
                if parser_name == 'общий':
                    continue
                    
                for row_num in range(2, len(df) + 2):
                    cell = ws.cell(row=row_num, column=status_col_idx)
                    status = cell.value
                    
                    if status == 'Найдено':
                        cell.fill = self.STATUS_FOUND_FILL
                    elif status == 'Цена по запросу':
                        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Оранжевый
                    elif status == 'Снят с производства':
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Серый
                    elif status in ['Товар не найден на сайте', 'Не найдено']:
                        cell.fill = self.STATUS_NOT_FOUND_FILL
                    elif status == 'Ошибка парсинга':
                        cell.fill = self.STATUS_ERROR_FILL
            
            # Автоподбор ширины колонок
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем изменения
            wb.save(file_path)
            wb.close()
            
            self.log.debug("formatting_parallel_applied", file=str(file_path))
            
        except Exception as e:
            self.log.warning("formatting_parallel_failed", error=str(e), error_type=type(e).__name__)

